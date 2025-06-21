import sqlite3
import openpyxl

# Caminhos dos ficheiros
excel_path = "resources/cdm/cdm_fazeconta.xlsx"
db_path = "db.sqlite3"

# Mapeamento simples dos tipos Django para SQLite
type_map = {
    "CharField": "TEXT",
    "TextField": "TEXT",
    "IntegerField": "INTEGER",
    "FloatField": "REAL",
    "BooleanField": "BOOLEAN",
    "DateField": "DATE",
    "DateTimeField": "DATETIME",
    "AutoField": "INTEGER PRIMARY KEY AUTOINCREMENT",
}

# Abre o Excel e a base de dados
wb = openpyxl.load_workbook(excel_path, data_only=True)
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Lê a tabela "Table Summary"
summary = wb["Table Summary"]
for row in summary.iter_rows(min_row=2, values_only=True):
    table_name, table_type = row
    if not table_name:
        continue
    sheet = wb[table_name]
    columns = []
    for col_row in sheet.iter_rows(min_row=2, values_only=True):
        column_name, django_field_type, datatype_parameters, auto_id, null_constraint = col_row
        # Tipo de dados
        sql_type = type_map.get(django_field_type, "TEXT")
        # Auto ID
        if auto_id and str(auto_id).lower() in ["yes", "true", "1"]:
            col_sql = f'"{column_name}" INTEGER PRIMARY KEY AUTOINCREMENT'
        else:
            col_sql = f'"{column_name}" {sql_type}'
            # NOT NULL ou NULL
            if null_constraint and null_constraint.upper() == "NOT NULL":
                col_sql += " NOT NULL"
            else:
                col_sql += " NULL"
        columns.append(col_sql)
    # Cria a tabela se não existir
    sql = f'CREATE TABLE IF NOT EXISTS "{table_name}" ({", ".join(columns)});'
    cursor.execute(sql)

conn.commit()
conn.close()
print("Tabelas criadas/atualizadas")